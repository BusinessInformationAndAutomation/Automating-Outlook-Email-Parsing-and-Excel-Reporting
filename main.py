import win32com.client
import re
from datetime import datetime, date, timedelta
from openpyxl import load_workbook

# Load Excel workbook path from text file
with open("excel_path.txt", 'r') as file:
    excel_ws = file.read().strip()

# Email subject prefix to filter
EMAIL_SUBJECT_PREFIX = "Departure for"

# Determine correct start date based on weekday
today = datetime.today()
if today.weekday() == 0:  # Monday
    start_date = today - timedelta(days=3)
else:
    start_date = today - timedelta(days=1)

search_date = start_date.strftime("%d-%m-%Y")
current_date = today.strftime("%d-%m-%Y")
filter_query = f"[ReceivedTime] >= '{search_date} 12:00 AM' AND [ReceivedTime] <= '{current_date} 11:59 PM'"

# Connect to Outlook
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
inbox = outlook.GetDefaultFolder(6)
emails = inbox.Items.Restrict(filter_query)

# Filter emails with matching subject
filtered_emails = [mail for mail in emails if mail.Subject and mail.Subject.startswith(EMAIL_SUBJECT_PREFIX)]

# Extract values from email body using regex
extracted_values = []
for email in filtered_emails:
    matches = re.findall(r"Departure check result: There are total of\s*(\d+)", email.Body, re.IGNORECASE)
    count = ''.join(matches).strip()
    extracted_values.append(count)

# Update Excel workbook
def update_excel_with_values(values, workbook_path):
    wb = load_workbook(workbook_path)
    for i, days_ago in enumerate([1, 2, 3]):
        date_str = (today - timedelta(days=days_ago)).strftime("%d-%m-%Y")
        if date_str in wb.sheetnames:
            sheet = wb[date_str]
            sheet['B22'] = f"There are total of {values[i]}"
    wb.save(workbook_path)

if today.weekday() == 0:
    update_excel_with_values(extracted_values, excel_ws)
